import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import *
import os, shutil
import string
import random
from tkinter import messagebox
import tempfile
import pyAesCrypt


class Register:

	def __init__( self, master=None ):
		# build ui
		frame_2 = ttk.Frame( master )
		user_L = ttk.Label( frame_2 )
		user_L.config( text='User Name' )
		user_L.grid( padx='5', pady='5' )
		self.user_E = ttk.Entry( frame_2 )
		self.user_E.config( font='{Arial} 12 {bold}', width='22' )
		self.user_E.grid( column='1', padx='5', pady='5', row='0' )
		password_L = ttk.Label( frame_2 )
		password_L.config( text='Password' )
		password_L.grid( column='0', padx='5', pady='5', row='1' )
		self.password_E = ttk.Entry( frame_2 )
		self.password_E.config( font='{Arial} 12 {bold}', width='22', show='•' )
		self.password_E.grid( column='1', padx='5', pady='5', row='1' )
		reenter_L = ttk.Label( frame_2 )
		reenter_L.config( text='Re Enter' )
		reenter_L.grid( column='0', padx='10', pady='5', row='2' )
		self.reenter_E = ttk.Entry( frame_2 )
		self.reenter_E.config( font='{Arial} 12 {bold}', width='22', show='•' )
		self.reenter_E.grid( column='1', padx='5', pady='5', row='2' )
		register_B = ttk.Button( frame_2 )
		register_B.config( default='normal', text='Register' )
		register_B.grid( column='0', padx='10', pady='20', row='3' )
		register_B.configure( command=self.register )
		cancle_B = ttk.Button( frame_2 )
		cancle_B.config( default='normal', text='Cancle' )
		cancle_B.grid( column='2', row='3' )
		cancle_B.configure( command=self.cancle )
		frame_2.config( height='200', width='200' )
		frame_2.place( anchor='nw', height='250', width='480', x='0', y='10' )

		# Main widget
		self.mainwindow = frame_2
		self.config()

	def config( self ):
		self.bufferSize = 64 * 1024
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		self.home = "Invisible"

	def random_str( self ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=12 ) ) )

	def register( self ):
		self.user = self.user_E.get()
		self.password = self.password_E.get()
		self.reenter = self.reenter_E.get()
		if ( self.user == "" or self.password == "" or self.reenter == "" ):
			messagebox.showerror( "Empty", "Empty Field !" )
		elif ( self.reenter_E.get() != self.password_E.get() ):
			messagebox.showerror( "Match", "Password Not Matched !" )
		else:
			self.userdir = os.path.join( self.home, 'Users' )
			if not os.path.exists( self.userdir ):
				os.makedirs( self.userdir )
			self.userxl = os.path.join( self.userdir, self.user )
			# self.tempxl = tempfile.mkstemp( suffix='.xlsx' )[ 1 ]
			self.tempxl = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )
			if not os.path.exists( self.userxl ):
				WB = Workbook()
				sheet = WB.active
				dirname = self.random_str()
				while os.path.exists( os.path.join( self.home, dirname ) ):
					dirname = self.random_str()
				dirlists = self.random_str()
				filelists = self.random_str()
				sheet.cell( row=1, column=1 ).value = self.user_E.get()
				sheet.cell( row=1, column=2 ).value = self.password_E.get()
				sheet.cell( row=1, column=3 ).value = dirname
				sheet.cell( row=1, column=4 ).value = dirlists
				sheet.cell( row=1, column=5 ).value = filelists

				WB.save( self.tempxl )

				pyAesCrypt.encryptFile( self.tempxl, self.userxl, self.password, self.bufferSize )
				WB.close()
				if not os.path.exists( os.path.join( self.home, dirname ) ):
					os.makedirs( os.path.join( self.home, dirname ) )

				if not os.path.exists( os.path.join( self.home, dirname, dirlists ) ):
					tempxl = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )
					WB = Workbook()
					sheet = WB.active
					sheet.cell( row=1, column=1 ).value = "SRC_DIR"
					sheet.cell( row=1, column=2 ).value = "TEMP DIR"

					WB.save( tempxl )
					pyAesCrypt.encryptFile(
					    tempxl, os.path.join( self.home, dirname, dirlists ), self.password, self.bufferSize
					    )
					WB.close()
					os.remove( tempxl )
				os.system( "attrib +s +h /s \"{}/*.*\"".format( self.userdir ) )
				# os.system( "attrib +s +h /s {}\\*.*".format( dirname ) )
				os.system( "attrib +s +h \"{}\"".format( self.userdir ) )
				os.system( "attrib +s +h /s \"{}/*.*\"".format( os.path.join( self.home, dirname ) ) )
				os.system( "attrib +s +h \"{}\"".format( os.path.join( self.home, dirname ) ) )

				messagebox.showinfo( "Done", "User {} Successfully Registerd !".format( self.user ) )
				self.mainwindow.destroy()
			else:
				messagebox.showerror( 'Found', "User With This Name Already Registerd !" )

			# pyAesCrypt.decryptFile( self.userxl, self.tempxl, "946dChPFx6j7IcR", self.bufferSize )
			# WB = load_workbook( self.tempxl )
			# sheet = WB.active
			# L_R = sheet.max_row
			# #print( L_R )
			# sheet.cell( row=L_R, column=1 ).value = self.user_E.get()
			# sheet.cell( row=L_R, column=2 ).value = self.password_E.get()
			# sheet.cell( row=L_R, column=3 ).value = self.random_str()
			# sheet.cell( row=L_R, column=4 ).value = self.random_str()
			# pyAesCrypt.encryptFile( self.tempxl, self.userxl, "946dChPFx6j7IcR", self.bufferSize )

			os.remove( self.tempxl )

	def cancle( self ):
		self.mainwindow.destroy()

	def run( self ):
		self.mainwindow.mainloop()


if __name__ == '__main__':
	root = tk.Tk()
	app = Register( root )
	#print( app.random_str() )
	# app.run()
