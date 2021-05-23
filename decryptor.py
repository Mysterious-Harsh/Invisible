import os
import string
import random
import shutil
import pyAesCrypt
import tempfile
from openpyxl import *


class Decryptor:

	def __init__( self, src_dir, dirlists, filelists, key ):
		self.key = key
		self.src_dir = src_dir
		# self.dest_dir = dest_dir
		self.bufferSize = 64 * 1024
		self.dirlists = dirlists
		self.filelists = filelists
		# password = "foopassword"

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def decrypt( self ):
		# D = os.path.dirname(__file__)
		# self.src_dir = os.path.join(D,"hello")
		# os.mkdir(self.src_dir)
		# self.src_dir = os.path.abspath(self.src_dir)
		# print(self.src_dir)
		FILEPATHXL = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
		# DIRPATHXL = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )
		# if os.path.exists( os.path.join( self.src_dir, self.dirlists ) ):
		# 	pyAesCrypt.decryptFile( os.path.join( self.src_dir, self.dirlists ), DIRPATHXL, self.key, self.bufferSize )
		# 	WB = load_workbook( DIRPATHXL )
		# 	sheet = WB.active
		# 	self.src_dir = sheet.cell( row=dest_dir, column=2 ).value
		# 	WB.close()
		# else:
		# 	print( "Invalid" )
		# print( self.src_dir )
		self.Flag = False

		if os.path.exists( os.path.join( self.src_dir, self.filelists ) ):
			pyAesCrypt.decryptFile(
			    os.path.join( self.src_dir, self.filelists ), FILEPATHXL, self.key, self.bufferSize
			    )
			WB = load_workbook( FILEPATHXL )
			sheet = WB.active
			M_R = sheet.max_row
		else:
			WB = Workbook()
			sheet = WB.active
			M_R = sheet.max_row
		self.Return = os.path.dirname( sheet.cell( row=1, column=1 ).value ).replace( ':', '' )
		# print( self.Return )
		for i in range( 1, M_R + 1 ):
			filepath = sheet.cell( row=i, column=1 ).value
			temp = sheet.cell( row=i, column=2 ).value
			DIR = os.path.dirname( filepath )
			# print( DIR )
			try:
				if not os.path.exists( DIR ):
					os.makedirs( DIR )
			except:
				# DIR = 'Forgotten' + "\\" + DIR.split( '\\' )[ -1 ]
				# DIR = 'Forgotten' + "\\" + '\\'.join( DIR.split( '\\' )[ 1 : ] )
				# print( DIR )
				DIR = DIR.replace( ':', '' )
				if not os.path.exists( DIR ):
					os.makedirs( DIR )
				filepath = os.path.join( DIR, os.path.basename( filepath ) )
				self.Flag = True

				# print( filepath )
			try:
				pyAesCrypt.decryptFile( temp, filepath, self.key, self.bufferSize )
			except:
				file, ext = os.path.splitext( filepath )
				file = file + "_" + self.random_str( 5 ) + ext
				pyAesCrypt.decryptFile( temp, file, self.key, self.bufferSize )

		for root, dirs, files in os.walk( self.src_dir ):
			for name in files:
				try:
					os.remove( os.path.join( root, name ) )
				except:
					pass
			for name in dirs:
				try:
					shutil.rmtree( os.path.join( root, name ) )
				except:
					pass
		try:
			shutil.rmtree( self.src_dir )
		except:
			pass
		os.remove( FILEPATHXL )
		# shutil.copy2( temp, filepath )

		if self.Flag:
			return self.Return
		else:
			return None


# D = Decryptor( "invisible", 1, "dirlist", "filelist", "hello" )
# D.decrypt()
