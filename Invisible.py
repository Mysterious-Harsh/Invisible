import tkinter as tk
import tkinter.ttk as ttk
import os
from register import Register
from tkinter import messagebox, PhotoImage
import tempfile
import pyAesCrypt
from openpyxl import *
from home_page import Homepage
import string
import random
from tkinter import END
from check_registration import check_registration
from Validate import Validation


class KeyApp:

	def __init__( self, master=None ):
		# build ui
		self.root = master
		self.root.title( "Invisible" )

		self.root.tk.call( "wm", "iconphoto", self.root._w, PhotoImage( file="images/logo.png" ) )
		self.keyframe = ttk.Labelframe( self.root )
		frame_1 = ttk.Frame( self.keyframe )
		self.getkey_MTV = tk.StringVar()
		self.getkey_M = tk.Message( frame_1 )
		self.getkey_M.config(
		    anchor='center',
		    background='#ffffff',
		    font='{Arial} 13 {bold}',
		    text='You Do not have a License Key.\nEnter Your Name.',
		    width='550',
		    textvariable=self.getkey_MTV
		    )
		self.getkey_MTV.set( 'You Do not have a License Key.\nEnter Your Name.' )
		self.getkey_M.pack( pady='10', side='top' )
		self.getkey_E = ttk.Entry( frame_1 )
		self.getkey_E.config( font='{Arial} 12 {bold}', width='26' )
		self.getkey_E.delete( '0', 'end' )
		# self.getkey_E.insert( '0', _text_ )
		self.getkey_E.pack( pady='10', side='top' )
		self.getkey_B = ttk.Button( frame_1 )
		self.getkey_B.config( text='Get Key' )
		self.getkey_B.pack( pady='10', side='top' )
		self.getkey_B.configure( command=self.get_key )
		frame_1.config( height='250', width='450' )
		frame_1.pack( anchor='n', expand='true', fill='both', pady='10', side='top' )
		frame_1.pack_propagate( 0 )
		self.keyframe.config( height='250', labelanchor='n', relief='flat', text='Invisible' )
		self.keyframe.config( width='480' )
		self.keyframe.pack( anchor='n', expand='true', fill='both', side='top' )
		self.keyframe.pack_propagate( 0 )
		self.root.config( height='200', width='200' )
		self.root.resizable( False, False )

		# Main widget
		self.mainwindow = self.root
		self.config()

	def config( self ):
		# self.filepath = ""
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		# self.home = "Invisible"
		# if os.path.exists( self.home ):
		# 	os.system( "attrib +s +h /s {}/*.*".format( self.home ) )
		# 	os.system( "attrib +s +h {}".format( self.home ) )
		# self.userdir = os.path.join( self.home, 'Users' )
		# self.bufferSize = 64 * 1024
		self.s_w = self.root.winfo_screenwidth()
		self.s_h = self.root.winfo_screenheight()
		self.wow = 480
		self.how = 250
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )
		# self.root.resizable( False, False )
		self.style = ttk.Style( self.mainwindow )
		self.style.configure( "TLabelframe", background="#ffffff", highlightthickness=5, relief="flat" )
		self.style.configure( "TFrame", background="#ffffff", highlightthickness=5, relief="flat" )
		self.style.configure(
		    "TLabelframe.Label",
		    background="#141414",
		    foreground="#ffffff",
		    font=( 'Arial Black', 16, 'bold' ),
		    width=self.wow,
		    anchor="center",
		    relief="flat"
		    )
		self.style.configure(
		    "Treeview",
		    fieldbackground="#ffffff",
		    background="#ffffff",
		    foreground="#000000",
		    font=( "Arial", 11, "bold" ),
		    borderwidth=5,
		    relief="flat"
		    )
		self.style.configure( "Treeview.Heading", font=( "Arial", 12, "bold" ) )

		self.style.configure(
		    "TButton",
		    background="#000000",
		    foreground="#000000",
		    font=( "Arial", 13, "bold" ),
		    relief="flat",
		    highlightcolor="blue",
		    bordercolor="#ffffff",
		    takefocus=True,
		    )
		# self.style.map( "TButton", background=[ ( 'active', 'blue' ) ] )
		self.style.configure(
		    "TLabel",
		    padding=0,
		    font=( "Arial", 14, "bold" ),
		    relief="flat",
		    width=10,
		    takefocus=False,
		    bordercolor="#ffffff",
		    background="#000000",
		    cursor="arrow",
		    anchor="center"
		    )

		self.style.configure( "TEntry", fieldbackground="#99ddff", font=( "Arial", 12, "bold" ), width='25' )
		self.count = 0

	def get_key( self ):
		name = self.getkey_E.get()
		if name == "":
			messagebox.showerror( 'Empty', "Please Enter Your Name !" )
		else:

			self.getkey_MTV.set( "Enter The Key." )
			self.getkey_B.configure( command=self.verify_Key, text="Verify Key" )
			self.getkey_E.delete( 0, "end" )
			self.Validate = Validation( name )

		pass

	def verify_Key( self ):
		key = self.getkey_E.get()
		if key == "":
			messagebox.showerror( 'Empty', "Please Enter Key !" )
		else:
			self.flag = self.Validate.compare( key )
			if self.flag:
				CK = check_registration()
				CK.Register( key )
				self.keyframe.destroy()
				app = InvisibleApp( self.root )
				app.run()
				# HP = HomePage( self.Mainframe, self.root )
				# self.homepage()
			else:
				messagebox.showerror( 'Wrong', "Wrong Key !" )
				self.count = self.count + 1
				if self.count == 3:
					self.root.destroy()

		pass

	def run( self ):
		self.mainwindow.mainloop()


class InvisibleApp:

	def __init__( self, master=None ):
		# build ui
		# self.root = tk.Toplevel(master)
		self.root = master
		self.root.title( "Invisible" )

		self.root.tk.call( "wm", "iconphoto", self.root._w, PhotoImage( file="images/logo.png" ) )
		self.mainframe = ttk.Labelframe( self.root )
		frame_1 = ttk.Frame( self.mainframe )
		user_L = ttk.Label( frame_1 )
		user_L.config( font='{Arial} 14 {bold}', takefocus=False, text='User Name' )
		user_L.grid( padx='10' )
		self.user_E = ttk.Entry( frame_1 )
		self.user_E.config( font='{Arial} 12 {bold}', takefocus=True, width='22' )
		self.user_E.grid( column='1', padx='5', pady='5', row='0' )
		password_L = ttk.Label( frame_1 )
		password_L.config( font='{Arial} 14 {bold}', takefocus=False, text='Password' )
		password_L.grid( column='0', padx='10', row='1' )
		self.password_E = ttk.Entry( frame_1 )
		self.password_E.config( font='{Arial} 12 {bold}', takefocus=True, width='22', show='•' )
		self.password_E.grid( column='1', padx='5', pady='5', row='1' )
		self.password_E.bind( '<Return>', self.login )
		login_B = ttk.Button( frame_1 )
		login_B.config( text='Login' )
		login_B.grid( column='0', pady='10', row='2' )
		login_B.configure( command=self.login )
		register_B = ttk.Button( frame_1 )
		register_B.config( takefocus=True, text='Register' )
		register_B.grid( column='2', pady='50', row='2' )
		register_B.configure( command=self.register )
		# changepass_B = ttk.Button( frame_1 )
		# changepass_B.config( text='Change Password' )
		# changepass_B.grid( column='2', pady='10', row='2' )
		frame_1.config( height='300', width='450' )
		frame_1.pack( anchor='n', expand='true', fill='both', side='top', pady='20' )
		frame_1.pack_propagate( 0 )
		self.mainframe.config( height='300', labelanchor='n', relief='flat', text='Invisible' )
		self.mainframe.config( width='480' )
		self.mainframe.pack( anchor='n', expand='true', fill='both', side='top' )
		self.mainframe.pack_propagate( 0 )
		# self.root.config( height='200', width='200' )
		# self.root.resizable( False, False )

		# Main widget
		self.mainwindow = self.root
		self.config()
		self.user_E.focus_set()

	def config( self ):
		self.filepath = ""
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		self.home = "Invisible"
		if os.path.exists( self.home ):
			os.system( "attrib +s +h /s \"{}/*.*\"".format( self.home ) )
			os.system( "attrib +s +h \"{}\"".format( self.home ) )

		#print( os.getcwd() )
		os.system( "attrib +s +h /s \"{}/*.*\"".format( os.getcwd() ) )
		os.system( "attrib +s +h \"{}\"".format( os.getcwd() ) )

		self.userdir = os.path.join( self.home, 'Users' )
		self.bufferSize = 64 * 1024
		self.s_w = self.root.winfo_screenwidth()
		self.s_h = self.root.winfo_screenheight()
		self.wow = 480
		self.how = 250
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )
		# self.root.resizable( False, False )
		self.style = ttk.Style( self.mainwindow )
		self.style.configure( "TLabelframe", background="#ffffff", highlightthickness=5, relief="flat" )
		self.style.configure( "TFrame", background="#ffffff", highlightthickness=5, relief="flat" )
		self.style.configure(
		    "TLabelframe.Label",
		    background="#141414",
		    foreground="#ffffff",
		    font=( 'Arial Black', 16, 'bold' ),
		    width=self.wow,
		    anchor="center",
		    relief="flat"
		    )
		self.style.configure(
		    "Treeview",
		    fieldbackground="#ffffff",
		    background="#ffffff",
		    foreground="#000000",
		    font=( "Arial", 11, "bold" ),
		    borderwidth=5,
		    relief="flat"
		    )
		self.style.configure( "Treeview.Heading", font=( "Arial", 12, "bold" ) )

		self.style.configure(
		    "TButton",
		    background="#000000",
		    foreground="#000000",
		    font=( "Arial", 13, "bold" ),
		    relief="flat",
		    highlightcolor="blue",
		    bordercolor="#ffffff",
		    takefocus=True,
		    )
		# self.style.map( "TButton", background=[ ( 'active', 'blue' ) ] )
		self.style.configure(
		    "TLabel",
		    padding=0,
		    font=( "Arial", 14, "bold" ),
		    relief="flat",
		    width=10,
		    takefocus=False,
		    bordercolor="#ffffff",
		    background="#ffffff",
		    cursor="arrow",
		    anchor="center"
		    )

		self.style.configure( "TEntry", fieldbackground="#99ddff", font=( "Arial", 12, "bold" ), width='22' )
		self.count = 0

	def random_str( self ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=12 ) ) )

	def register( self ):
		if not os.path.exists( self.home ):
			os.makedirs( self.home )
			os.system( "attrib +s +h \"{}\"".format( self.home ) )
		#print( self.home )
		Re = Register( self.mainframe )

		pass

	def login( self, event=None ):
		self.user = self.user_E.get()
		self.password = self.password_E.get()
		self.userxl = os.path.join( self.userdir, self.user )
		if ( self.user == "" or self.password == "" ):
			messagebox.showerror( "Empty", "Empty Field !" )
		elif not os.path.exists( self.userxl ):
			messagebox.showerror( 'Invalid', "Invalid User Name !" )
		else:
			self.tempxl = os.path.join( tempfile.gettempdir(), self.random_str() + '.xlsx' )
			try:
				pyAesCrypt.decryptFile( self.userxl, self.tempxl, self.password, self.bufferSize )
				WB = load_workbook( self.tempxl )
				sheet = WB.active
				self.dirname = sheet.cell( row=1, column=3 ).value
				self.dirlists = sheet.cell( row=1, column=4 ).value
				self.filelists = sheet.cell( row=1, column=5 ).value
				os.remove( self.tempxl )
				#print( self.user, self.password, self.dirname, self.dirlists, self.filelists )
				self.user_E.delete( 0, END )
				self.password_E.delete( 0, END )
				hp = Homepage(
				    self.root, self.mainframe, self.user, self.password, self.dirname, self.dirlists, self.filelists
				    )

			except Exception as e:
				print( e )
				messagebox.showerror( 'Wrong', "Wrong Password !" )

		pass

	def run( self ):
		self.mainwindow.mainloop()


if __name__ == '__main__':
	CR = check_registration()
	Flag = CR.check()
	root = tk.Tk()
	if Flag:
		app = InvisibleApp( root )
		app.run()
	else:
		app = KeyApp( root )
		app.run()
