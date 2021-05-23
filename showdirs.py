import tkinter as tk
import tkinter.ttk as ttk
from encryptor import Encryptor
from decryptor import Decryptor
from tkinter import messagebox, filedialog
import os, shutil
import random
import string
import pyAesCrypt
from openpyxl import *
import tempfile
from tkinter import *


class Showdirs:

	def __init__( self, root, master, home, user, password, dirname, dirlists, filelists ):
		# build ui
		self.root = root
		self.mainframe = master
		self.user = user
		self.password = password
		self.dirname = dirname
		self.dirlists = dirlists
		self.filelists = filelists
		self.home = home
		# master = root
		self.show_F = ttk.Frame( master )
		self.dirlist_LB = tk.Listbox( self.show_F )
		self.dirlist_LB.config( font='{Arial} 12 {bold}', height='12', relief='flat', selectmode='single' )
		self.dirlist_LB.config( state='normal', takefocus=True, width='10' )
		# self.dirlist_LB.place( anchor='nw', x='0', y='0' )
		self.dirlist_LB.pack( expand='true', fill='both', side='left' )
		self.scrollbar_V = ttk.Scrollbar( self.show_F )
		self.scrollbar_V.config( orient='vertical' )
		self.scrollbar_V.pack( expand='true', fill='y', side='top' )
		self.scrollbar_V.config( command=self.dirlist_LB.yview )
		self.dirlist_LB.configure( yscrollcommand=self.scrollbar_V.set )
		open_B = ttk.Button( self.show_F )
		open_B.config( text='Open' )
		open_B.place( anchor='n', relx='0.12', rely='0.92', x='0', y='0' )
		open_B.configure( command=self.open )
		hide_B = ttk.Button( self.show_F )
		hide_B.config( text='Hide' )
		hide_B.place( anchor='n', relx='0.36', rely='0.92', x='0', y='0' )
		hide_B.configure( command=self.hide )
		Delete_B = ttk.Button( self.show_F )
		Delete_B.config( text='Delete' )
		Delete_B.place( anchor='n', relx='0.61', rely='0.92', x='0', y='0' )
		Delete_B.configure( command=self.delete )
		close_B = ttk.Button( self.show_F )
		close_B.config( text='Close' )
		close_B.place( anchor='n', relx='0.85', rely='0.92', x='0', y='0' )
		close_B.configure( command=self.close )
		self.show_F.config( height='500', relief='flat', width='480' )
		self.show_F.place( anchor='nw', height='500', width='480', x='0', y='0' )

		# Main widget
		self.mainwindow = self.show_F
		# for line in range( 100 ):
		# 	self.dirlist_LB.insert( END, "This is line number " + str( line ) )
		self.config()
		self.insert()
		self.dirlist_LB.focus_set()
		#print( self.home )

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def config( self ):
		# self.filepath = ""
		# self.home = os.path.abspath( os.path.join( os.path.expanduser( '~' ), "AppData\\Local\\Programs\\Invisible" ) )
		# self.userdir = os.path.join( self.home, 'Users' )
		self.bufferSize = 64 * 1024
		self.s_w = self.root.winfo_screenwidth()
		self.s_h = self.root.winfo_screenheight()
		self.wow = 480
		self.how = 530
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )

	def insert( self ):
		self.dirlist_LB.delete( 0, END )
		tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
		self.dirlist = []
		if os.path.exists( os.path.join( self.home, self.dirname, self.dirlists ) ):
			pyAesCrypt.decryptFile(
			    os.path.join( self.home, self.dirname, self.dirlists ), tempxl, self.password, self.bufferSize
			    )
			WB = load_workbook( tempxl )
			sheet = WB.active
			L_R = sheet.max_row
			#print( L_R )
			for i in range( 2, L_R + 1 ):
				self.dirlist_LB.insert( END, os.path.basename( sheet.cell( row=i, column=1 ).value ) )
				self.dirlist.append( [ sheet.cell( row=i, column=1 ).value, sheet.cell( row=i, column=2 ).value ] )
				#print( sheet.cell( row=i, column=1 ).value )
			# root_dir = sheet.cell( row=dest_dir, column=2 ).value
			WB.close()
			os.remove( tempxl )
		self.dirlist_LB.selection_set( 0 )
		#print( self.dirlist )

	def hide( self ):
		selected = self.dirlist_LB.curselection()
		# self.src_dir = filedialog.askdirectory( title="Select Folder" )
		if selected != ():
			src = self.dirlist[ selected[ 0 ] ][ 0 ]
			if os.path.exists( src ):
				dest = os.path.join( self.home, self.dirname )
				self.mainframe.config( text="Hidding..." )
				self.mainframe.update()
				# self.src_dir = os.path.abspath( self.src_dir )
				Enc = Encryptor( src, dest, self.dirlists, self.filelists, self.password )
				Enc.encrypt()
				self.mainframe.config( text=self.user )
				self.mainframe.update()
				self.insert()
			else:
				messagebox.showerror( 'Hidden', 'Folder allready hidden !' )

		else:
			messagebox.showerror( 'Error', "Folder not Selected !" )

	def open( self ):
		selected = self.dirlist_LB.curselection()
		if selected != ():
			self.mainframe.config( text="Opening..." )
			self.mainframe.update()
			src = self.dirlist[ selected[ 0 ] ][ 1 ]
			dest = self.dirlist[ selected[ 0 ] ][ 0 ]
			if os.path.exists( src ):
				#print( selected[ 0 ] )
				Dc = Decryptor( src, self.dirlists, self.filelists, self.password )
				NewPath = Dc.decrypt()
				if NewPath != None:
					dest = self.dirlist[ selected[ 0 ] ][ 0 ] = NewPath

				os.startfile( dest )

			else:
				if os.path.exists( dest ):
					messagebox.showinfo( 'Info', "Folder Already Opened !" )
					os.startfile( dest )
				else:
					flag = "no"
					flag = messagebox.askquestion( 'Info', "Folder not found. You want to remove it from list ?" )
					if flag == "yes":
						tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
						pyAesCrypt.decryptFile(
						    os.path.join( self.home, self.dirname, self.dirlists ), tempxl, self.password,
						    self.bufferSize
						    )
						WB = load_workbook( tempxl )
						sheet = WB.active
						sheet.delete_rows( selected[ 0 ] + 2 )
						WB.save( tempxl )
						WB.close()
						os.system(
						    "attrib -s -h  \"{}\"".format( os.path.join( self.home, self.dirname, self.dirlists ) )
						    )
						pyAesCrypt.encryptFile(
						    tempxl, os.path.join( self.home, self.dirname, self.dirlists ), self.password,
						    self.bufferSize
						    )
						os.system(
						    "attrib +s +h  \"{}\"".format( os.path.join( self.home, self.dirname, self.dirlists ) )
						    )
						os.remove( tempxl )
						self.insert()
			self.mainframe.config( text=self.user )
			self.mainframe.update()

		else:
			messagebox.showerror( 'Invalid', "Invalid Selection !" )

	def delete( self ):
		selected = self.dirlist_LB.curselection()
		if selected != ():
			src = self.dirlist[ selected[ 0 ] ][ 1 ]
			dest = self.dirlist[ selected[ 0 ] ][ 0 ]

			flag = 'yes'
			flag = messagebox.askquestion( "Open", "You want to open this folder ?" )
			if flag == 'yes':
				self.mainframe.config( text="Opening..." )
				self.mainframe.update()
				if os.path.exists( src ):
					#print( selected[ 0 ] )
					Dc = Decryptor( src, self.dirlists, self.filelists, self.password )
					NewPath = Dc.decrypt()
					if NewPath != None:
						dest = self.dirlist[ selected[ 0 ] ][ 0 ] = NewPath
					os.startfile( dest )

				else:
					if os.path.exists( dest ):
						messagebox.showinfo( 'Info', "Folder Already Opened !" )
						os.startfile( dest )
					else:
						messagebox.showinfo( 'Info', "Folder Already Deleted !" )
				self.mainframe.config( text=self.user )
				self.mainframe.update()

			elif flag == "no":
				flag = 'no'
				flag = messagebox.askquestion( "Hide", "Are you sure you want to permanently delete this folder ?" )
				if flag == 'yes':
					self.mainframe.config( text="Deleting..." )
					self.mainframe.update()
					if os.path.exists( src ):
						for root, dirs, files in os.walk( src ):
							for name in files:
								try:
									os.remove( os.path.join( root, name ) )
								except:
									pass
							for name in dirs:
								try:
									shutil.rmtree( os.path.join( root, name ) )
								except:
									pass
						try:
							shutil.rmtree( src )
						except:
							pass

					else:
						messagebox.showinfo( 'Info', "Folder Already Deleted !" )
						# os.startfile( dest )
					self.mainframe.config( text=self.user )
					self.mainframe.update()
			if flag == 'yes':

				tempxl = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
				pyAesCrypt.decryptFile(
				    os.path.join( self.home, self.dirname, self.dirlists ), tempxl, self.password, self.bufferSize
				    )
				WB = load_workbook( tempxl )
				sheet = WB.active
				sheet.delete_rows( selected[ 0 ] + 2 )
				WB.save( tempxl )
				WB.close()
				os.system( "attrib -s -h  \"{}\"".format( os.path.join( self.home, self.dirname, self.dirlists ) ) )
				pyAesCrypt.encryptFile(
				    tempxl, os.path.join( self.home, self.dirname, self.dirlists ), self.password, self.bufferSize
				    )
				os.system( "attrib +s +h  \"{}\"".format( os.path.join( self.home, self.dirname, self.dirlists ) ) )

				os.remove( tempxl )
				self.insert()

			#print( flag )

		else:
			messagebox.showerror( 'Invalid', "Invalid Selection !" )

	def close( self ):
		self.show_F.destroy()
		self.wow = 480
		self.how = 250
		x_c = ( ( self.s_w / 2 ) - ( self.wow / 2 ) )
		y_c = ( ( self.s_h / 2 ) - ( self.how / 2 ) )
		self.root.geometry( "%dx%d+%d+%d" % ( self.wow, self.how, x_c, y_c ) )

	def run( self ):
		self.mainwindow.mainloop()


if __name__ == '__main__':
	root = tk.Tk()
	app = Showdirs( root )
	app.run()
