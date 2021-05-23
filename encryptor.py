import os
import string
import random
import shutil
import pyAesCrypt
from openpyxl import *
import tempfile
from tkinter import messagebox


class Encryptor:

	def __init__( self, src_dir, dest_dir, dirlists, filelists, key ):
		self.key = key
		self.src_dir = src_dir
		self.dest_dir = dest_dir
		self.dirlists = dirlists
		self.filelists = filelists
		self.bufferSize = 64 * 1024

	def random_str( self, n ):
		return ( ''.join( random.choices( string.ascii_uppercase + string.digits + string.ascii_lowercase, k=n ) ) )

	def encrypt( self ):
		if not os.path.exists( self.dest_dir ):
			os.makedirs( self.dest_dir )

		root_dir = os.path.join( self.dest_dir, self.random_str( 12 ) )
		while os.path.exists( root_dir ):
			root_dir = os.path.join( self.dest_dir, self.random_str( 12 ) )

		os.makedirs( root_dir )
		# root_dir = os.path.abspath( root_dir )
		#print( root_dir )

		FILEPATHXL = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )
		DIRPATHXL = os.path.join( tempfile.gettempdir(), self.random_str( 12 ) + '.xlsx' )

		if os.path.exists( os.path.join( self.dest_dir, self.dirlists ) ):
			pyAesCrypt.decryptFile( os.path.join( self.dest_dir, self.dirlists ), DIRPATHXL, self.key, self.bufferSize )
			WB = load_workbook( DIRPATHXL )
			sheet = WB.active
		else:
			WB = Workbook()
			sheet = WB.active
			sheet.cell( row=1, column=1 ).value = "SRC_DIR"
			sheet.cell( row=1, column=2 ).value = "TEMP DIR"

		#print( DIRPATHXL )

		L_R = sheet.max_row + 1
		Flag = True
		close = True
		for i in range( 2, L_R ):
			if sheet.cell( row=i, column=1 ).value == self.src_dir:
				if not os.path.exists( sheet.cell( row=i, column=2 ).value ):
					sheet.cell( row=i, column=2 ).value = root_dir
				else:
					try:
						shutil.rmtree( root_dir )
					except:
						pass
					messagebox.showerror( 'Exist', "Please Rename This Folder !" )
					close = False
				Flag = False
		if close:
			#print( L_R )
			if Flag:
				sheet.cell( row=L_R, column=1 ).value = self.src_dir
				sheet.cell( row=L_R, column=2 ).value = root_dir
			WB.save( DIRPATHXL )
			# #print( ( self.dest_dir + "\\" + self.dirlists ) )
			os.system( "attrib -s -h  \"{}\"".format( self.dest_dir + "\\" + self.dirlists ) )
			pyAesCrypt.encryptFile( DIRPATHXL, os.path.join( self.dest_dir, self.dirlists ), self.key, self.bufferSize )
			WB.close()
			os.remove( DIRPATHXL )

			WB = Workbook()
			sheet = WB.active
			cell = 1
			for path, subdirs, files in os.walk( self.src_dir ):
				# for sub in subdirs:
				#     #print(os.path.abspath(sub))
				for name in files:
					temp = os.path.join( root_dir, self.random_str( 15 ) )
					while os.path.exists( temp ):
						temp = os.path.join( root_dir, self.random_str( 15 ) )
					filepath = os.path.abspath( os.path.join( path, name ) )
					# #print( filepath )
					# #print( temp )
					sheet.cell( row=cell, column=1 ).value = filepath
					sheet.cell( row=cell, column=2 ).value = temp
					pyAesCrypt.encryptFile( filepath, temp, self.key, self.bufferSize )
					# shutil.copy2(filepath, temp)
					cell += 1
			for root, dirs, files in os.walk( self.src_dir ):
				for name in files:
					try:
						os.remove( os.path.join( root, name ) )
					except:
						pass
				for name in dirs:
					try:
						shutil.rmtree( os.path.join( root, name ) )
					except:
						pass
			try:
				shutil.rmtree( self.src_dir )
			except:
				pass
			WB.save( FILEPATHXL )
			pyAesCrypt.encryptFile( FILEPATHXL, os.path.join( root_dir, self.filelists ), self.key, self.bufferSize )
			WB.close()
			os.remove( FILEPATHXL )
			os.system( "attrib +s +h /s \"{}/*.*\"".format( self.dest_dir ) )
			os.system( "attrib +s +h  \"{}\"".format( root_dir ) )
			os.system( "attrib +s +h  \"{}\"".format( self.dest_dir ) )

			messagebox.showinfo( 'Done', 'Folder Hidden !' )


# E = Encryptor( "test", 'invisible', "dirlist", "filelist", "hello" )
# E.encrypt()
